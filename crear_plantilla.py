
#!/usr/bin/env python3
"""
BARBER\u00cdA EL COLOCHO \u2014 Generador de Plantilla Excel
Ejecutar: python3 crear_plantilla.py
"""

import subprocess
import sys

def install(pkg):
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', pkg, '-q'])

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    install('openpyxl')
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from datetime import datetime, date
import os

# \u2500\u2500 Paleta de colores \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
GOLD       = "C8A951"
GOLD_LIGHT = "F0D080"
DARK_BG    = "1A1A2E"
DARK2_BG   = "16213E"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F8F9FA"
MEDIUM_GRAY= "E9ECEF"
SUCCESS    = "2ECC71"
WARNING    = "F39C12"
DANGER     = "E74C3C"
INFO       = "3498DB"
TEXT_DARK  = "2D3436"
TEXT_MUTED = "95A5A6"

# \u2500\u2500 Estilos base \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=TEXT_DARK, italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin():
    s = Side(style="thin", color="E0E0E0")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color=GOLD)
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg=DARK_BG, fg=GOLD, size=11):
    cell.fill    = fill(bg)
    cell.font    = font(bold=True, size=size, color=fg)
    cell.alignment = align("center")
    cell.border  = border_thin()

def style_subheader(cell, bg=GOLD, fg=DARK_BG):
    cell.fill    = fill(bg)
    cell.font    = font(bold=True, size=10, color=fg)
    cell.alignment = align("center")
    cell.border  = border_thin()

def style_data(cell, bg=WHITE):
    cell.fill    = fill(bg)
    cell.font    = font(size=10)
    cell.alignment = align()
    cell.border  = border_thin()

def style_money(cell, bg=WHITE):
    cell.fill    = fill(bg)
    cell.font    = font(size=10, color="9A7A2E", bold=True)
    cell.alignment = align("right")
    cell.border  = border_thin()
    cell.number_format = '"Q"#,##0.00'

def style_total(cell):
    cell.fill    = fill(DARK_BG)
    cell.font    = font(bold=True, size=12, color=GOLD)
    cell.alignment = align("right")
    cell.border  = border_medium()
    cell.number_format = '"Q"#,##0.00'

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_title(ws, cell_range, text, bg=DARK_BG, fg=GOLD, size=14):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill  = fill(bg)
    cell.font  = font(bold=True, size=size, color=fg)
    cell.alignment = align("center")

# \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
# CREAR WORKBOOK
# \u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550\u2550
wb = Workbook()

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 1: PORTADA
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_cover = wb.active
ws_cover.title = "\ud83d\udccb Inicio"
ws_cover.sheet_view.showGridLines = False

# Fondo oscuro completo
for row in ws_cover.iter_rows(min_row=1, max_row=40, min_col=1, max_col=10):
    for cell in row:
        cell.fill = fill(DARK_BG)

# Logo / T\u00edtulo
ws_cover.row_dimensions[1].height = 20
ws_cover.row_dimensions[2].height = 20
ws_cover.row_dimensions[3].height = 60
ws_cover.row_dimensions[4].height = 40
ws_cover.row_dimensions[5].height = 30
ws_cover.row_dimensions[6].height = 20
ws_cover.row_dimensions[7].height = 30

ws_cover.merge_cells("B3:I3")
c = ws_cover["B3"]
c.value = "\u2702  BARBER\u00cdA EL COLOCHO"
c.fill  = fill(DARK_BG)
c.font  = Font(name="Calibri", bold=True, size=36, color=GOLD)
c.alignment = align("center")

ws_cover.merge_cells("B4:I4")
c = ws_cover["B4"]
c.value = "Sistema de Gesti\u00f3n \u2014 Base de Datos Excel"
c.fill  = fill(DARK_BG)
c.font  = Font(name="Calibri", size=16, color=GOLD_LIGHT, italic=True)
c.alignment = align("center")

ws_cover.merge_cells("B5:I5")
c = ws_cover["B5"]
c.value = f"Generado el: {datetime.now().strftime('%d/%m/%Y  %H:%M')}"
c.fill  = fill(DARK_BG)
c.font  = Font(name="Calibri", size=11, color=TEXT_MUTED)
c.alignment = align("center")

# L\u00ednea dorada separadora
for col in range(2, 10):
    cell = ws_cover.cell(row=7, column=col)
    cell.fill = fill(GOLD)
    cell.value = ""
ws_cover.row_dimensions[7].height = 4

# Tabla de hojas disponibles
headers_info = ["Hoja", "Descripci\u00f3n", "Uso"]
data_info = [
    ["\ud83d\udcca Ventas",           "Registro de todas las ventas procesadas",    "Autom\u00e1tico al exportar"],
    ["\ud83d\uded2 Detalle Items",    "Desglose de cada producto por venta",         "Autom\u00e1tico al exportar"],
    ["\ud83d\udce6 Productos",        "Cat\u00e1logo de productos con precios y stock",   "Referencia / Actualizar"],
    ["\u2702\ufe0f Servicios",        "Servicios ofrecidos con precios",             "Referencia"],
    ["\ud83d\udc65 Clientes",         "Base de datos de clientes",                   "Autom\u00e1tico / Manual"],
    ["\ud83d\udc88 Barberos",         "Equipo de barberos y comisiones",             "Manual"],
    ["\ud83d\udcc5 Citas",            "Registro de citas y reservaciones",           "Manual / Autom\u00e1tico"],
    ["\ud83d\udcc8 Resumen Mensual",  "Gr\u00e1fica de ganancias por mes",                "An\u00e1lisis / Dashboard"],
    ["\ud83d\udcb0 Registro Ventas",  "Hoja de ingreso manual de ventas",            "Entrada de datos"],
]

start_row = 9
ws_cover.merge_cells(f"B{start_row}:I{start_row}")
c = ws_cover.cell(row=start_row, column=2)
c.value = "CONTENIDO DEL LIBRO"
c.fill = fill(GOLD)
c.font = Font(name="Calibri", bold=True, size=13, color=DARK_BG)
c.alignment = align("center")
ws_cover.row_dimensions[start_row].height = 26

for j, h in enumerate(headers_info, 2):
    c = ws_cover.cell(row=start_row+1, column=j)
    style_header(c, bg=DARK2_BG, fg=GOLD_LIGHT)
    c.value = h
ws_cover.row_dimensions[start_row+1].height = 20

for i, row_data in enumerate(data_info):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 2):
        c = ws_cover.cell(row=start_row+2+i, column=j)
        c.value = val
        c.fill = fill(bg)
        c.font = Font(name="Calibri", size=10, color=TEXT_DARK)
        c.alignment = align()
        c.border = border_thin()
    ws_cover.row_dimensions[start_row+2+i].height = 20

set_col_width(ws_cover, 2, 22)
set_col_width(ws_cover, 3, 45)
set_col_width(ws_cover, 4, 28)

# Footer
footer_row = start_row + 2 + len(data_info) + 2
ws_cover.merge_cells(f"B{footer_row}:I{footer_row}")
c = ws_cover.cell(row=footer_row, column=2)
c.value = "\u2702\ufe0f  Barber\u00eda El Colocho  \u2022  Tu look, nuestra pasi\u00f3n  \u2022  Sistema de Gesti\u00f3n v1.0"
c.fill = fill(DARK_BG)
c.font = Font(name="Calibri", size=10, color=TEXT_MUTED, italic=True)
c.alignment = align("center")

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 2: VENTAS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_v = wb.create_sheet("\ud83d\udcca Ventas")
ws_v.sheet_view.showGridLines = False
ws_v.freeze_panes = "A3"

merge_title(ws_v, "A1:M1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 REGISTRO DE VENTAS", size=13)
ws_v.row_dimensions[1].height = 28

col_headers_v = [
    ("#", 6), ("Fecha", 13), ("Hora", 8), ("Cliente", 22),
    ("Barbero", 18), ("Productos/Servicios", 40), ("Subtotal", 13),
    ("Desc%", 8), ("IVA", 10), ("Total", 13),
    ("M\u00e9todo Pago", 14), ("Nota", 22), ("Estado", 12)
]
for col, (h, w) in enumerate(col_headers_v, 1):
    c = ws_v.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_v.column_dimensions[get_column_letter(col)].width = w
ws_v.row_dimensions[2].height = 22

# Datos de ejemplo
ventas_ejemplo = [
    (1, "2025-06-15", "09:30", "Juan Garc\u00eda",    "Carlos P\u00e9rez",   "Corte Cl\u00e1sico x1",               50.00, 0, 0.00,  50.00, "Efectivo",      "",                    "completada"),
    (2, "2025-06-15", "10:45", "Pedro L\u00f3pez",    "Jos\u00e9 Mart\u00ednez",  "Corte + Barba x1 | Pomada x1",   125.00, 5, 0.00, 118.75, "Efectivo",      "Cliente frecuente",   "completada"),
    (3, "2025-06-15", "11:20", "Luis Torres",    "Carlos P\u00e9rez",   "Degradado/Fade x1 | Refresco x2", 89.00, 0, 0.00,  89.00, "Tarjeta",       "",                    "completada"),
    (4, "2025-06-16", "09:00", "Roberto Ruiz",   "Miguel R.",      "Tinte/Color x1",                 120.00, 10,0.00, 108.00, "Transferencia", "Tinte oscuro",        "completada"),
    (5, "2025-06-16", "14:30", "Consumidor",     "Jos\u00e9 Mart\u00ednez",  "Arreglo de Barba x1 | Aceite x1", 90.00, 0, 0.00,  90.00, "Efectivo",      "",                    "completada"),
]
for i, row_data in enumerate(ventas_ejemplo):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_v.cell(row=3+i, column=j)
        if j in (7, 9, 10):  # money columns
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
    ws_v.row_dimensions[3+i].height = 18

# Fila de total
total_row = 3 + len(ventas_ejemplo) + 1
ws_v.merge_cells(f"A{total_row}:F{total_row}")
c = ws_v.cell(row=total_row, column=1)
c.value = "TOTAL GENERAL"
c.fill = fill(DARK_BG)
c.font = Font(name="Calibri", bold=True, size=11, color=GOLD)
c.alignment = align("right")

c_total = ws_v.cell(row=total_row, column=10)
style_total(c_total)
c_total.value = sum(r[9] for r in ventas_ejemplo)
ws_v.row_dimensions[total_row].height = 24

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 3: DETALLE ITEMS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_di = wb.create_sheet("\ud83d\uded2 Detalle Items")
ws_di.sheet_view.showGridLines = False
ws_di.freeze_panes = "A3"

merge_title(ws_di, "A1:H1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 DETALLE DE ITEMS POR VENTA", size=12)
ws_di.row_dimensions[1].height = 26

col_di = [
    ("Venta#", 9), ("Fecha", 13), ("Cliente", 22), ("Tipo", 12),
    ("Producto/Servicio", 30), ("Precio Unit.", 14), ("Cantidad", 10), ("Total", 14)
]
for col, (h, w) in enumerate(col_di, 1):
    c = ws_di.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_di.column_dimensions[get_column_letter(col)].width = w
ws_di.row_dimensions[2].height = 20

items_ejemplo = [
    (1, "2025-06-15", "Juan Garc\u00eda",   "Servicio",  "Corte Cl\u00e1sico",      50.00, 1, 50.00),
    (2, "2025-06-15", "Pedro L\u00f3pez",   "Servicio",  "Corte + Barba",      80.00, 1, 80.00),
    (2, "2025-06-15", "Pedro L\u00f3pez",   "Producto",  "Pomada Capilar",     45.00, 1, 45.00),
    (3, "2025-06-15", "Luis Torres",   "Servicio",  "Degradado/Fade",     65.00, 1, 65.00),
    (3, "2025-06-15", "Luis Torres",   "Producto",  "Refresco Coca-Cola", 12.00, 2, 24.00),
    (4, "2025-06-16", "Roberto Ruiz",  "Servicio",  "Tinte/Color",       120.00, 1,120.00),
    (5, "2025-06-16", "Consumidor",    "Servicio",  "Arreglo de Barba",   35.00, 1, 35.00),
    (5, "2025-06-16", "Consumidor",    "Producto",  "Aceite de Barba",    55.00, 1, 55.00),
]
for i, row_data in enumerate(items_ejemplo):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_di.cell(row=3+i, column=j)
        if j in (6, 8):
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
        if j == 4:
            c.fill = fill("EBF5FB" if val == "Servicio" else "EAFAF1")
            c.font = Font(name="Calibri", size=10,
                          color=INFO if val == "Servicio" else SUCCESS, bold=True)
    ws_di.row_dimensions[3+i].height = 18

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 4: PRODUCTOS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_p = wb.create_sheet("\ud83d\udce6 Productos")
ws_p.sheet_view.showGridLines = False
ws_p.freeze_panes = "A3"

merge_title(ws_p, "A1:J1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 CAT\u00c1LOGO DE PRODUCTOS", size=12)
ws_p.row_dimensions[1].height = 26

col_p = [
    ("Nombre", 28), ("Categor\u00eda", 14), ("Marca", 14), ("Precio Venta", 14),
    ("Precio Costo", 14), ("Stock", 8), ("Stock M\u00edn.", 10),
    ("Estado", 12), ("Valor Inv.", 13), ("Descripci\u00f3n", 35)
]
for col, (h, w) in enumerate(col_p, 1):
    c = ws_p.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_p.column_dimensions[get_column_letter(col)].width = w
ws_p.row_dimensions[2].height = 20

productos = [
    ("Pomada Capilar Fuerte",  "Cabello",    "BarberPro",  45, 25, 20, 5,  "\u2705 Normal",  900,  "Fijaci\u00f3n fuerte con brillo"),
    ("Aceite de Barba",        "Barba",      "BeardKing",  55, 28, 15, 4,  "\u2705 Normal",  825,  "Hidrataci\u00f3n y suavizado de barba"),
    ("Crema para Afeitar",     "Cremas",     "Gillette",   38, 18, 12, 5,  "\u2705 Normal",  456,  "Crema de afeitado premium"),
    ("Loci\u00f3n Aftershave",      "Lociones",   "Old Spice",  60, 30, 10, 4,  "\u2705 Normal",  600,  "Loci\u00f3n post afeitado"),
    ("Shampoo para Hombre",    "Cabello",    "Head&Sh",    42, 20, 18, 5,  "\u2705 Normal",  756,  "Shampoo anticaspa"),
    ("Refresco Coca-Cola",     "Bebidas",    "Coca-Cola",  12,  7, 24,10,  "\u2705 Normal",  288,  "Refresco 500ml"),
    ("Agua Pura",              "Bebidas",    "Fuente",      8,  4, 30,10,  "\u2705 Normal",  240,  "Agua pura 500ml"),
    ("Gorra El Colocho",       "Gorras",     "Colocho",    85, 40,  8, 3,  "\u2705 Normal",  680,  "Gorra estampada marca propia"),
    ("Cera Mate",              "Cabello",    "Mandom",     50, 25, 14, 5,  "\u2705 Normal",  700,  "Cera mate fijaci\u00f3n media"),
    ("Loci\u00f3n Hidratante",      "Lociones",   "Nivea",      70, 35,  6, 4,  "\u26a0\ufe0f Bajo",   420,  "Loci\u00f3n hidratante corporal"),
    ("Peine Profesional",      "Accesorios", "Generic",    25, 10, 20, 5,  "\u2705 Normal",  500,  "Peine de bolsillo"),
    ("Gorra Snapback",         "Gorras",     "NewEra",    110, 55,  5, 3,  "\u26a0\ufe0f Bajo",   550,  "Gorra snapback ajustable"),
]
for i, row_data in enumerate(productos):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_p.cell(row=3+i, column=j)
        if j in (4, 5, 9):
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
        if j == 8:
            if "\u26a0\ufe0f" in str(val):
                c.fill = fill("FEF9E7")
                c.font = Font(name="Calibri", size=10, color=WARNING, bold=True)
            elif "\u274c" in str(val):
                c.fill = fill("FDEDEC")
                c.font = Font(name="Calibri", size=10, color=DANGER, bold=True)
            else:
                c.fill = fill("EAFAF1")
                c.font = Font(name="Calibri", size=10, color=SUCCESS, bold=True)
    ws_p.row_dimensions[3+i].height = 18

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 5: SERVICIOS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_s = wb.create_sheet("\u2702\ufe0f Servicios")
ws_s.sheet_view.showGridLines = False
ws_s.freeze_panes = "A3"

merge_title(ws_s, "A1:E1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 CAT\u00c1LOGO DE SERVICIOS", size=12)
ws_s.row_dimensions[1].height = 26

col_s = [("Nombre del Servicio", 30), ("Precio", 12), ("Duraci\u00f3n (min)", 16), ("Icono", 8), ("Descripci\u00f3n", 40)]
for col, (h, w) in enumerate(col_s, 1):
    c = ws_s.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_s.column_dimensions[get_column_letter(col)].width = w
ws_s.row_dimensions[2].height = 20

servicios = [
    ("Corte Cl\u00e1sico",         50,  30, "\u2702\ufe0f",  "Corte tradicional a tijera o m\u00e1quina"),
    ("Corte + Barba",         80,  45, "\ud83d\udc88",  "Corte y arreglo de barba"),
    ("Degradado / Fade",      65,  40, "\ud83e\ude92",  "Degradado de piel a n\u00famero"),
    ("Arreglo de Barba",      35,  20, "\ud83e\uddd4",  "Perfilado y arreglo de barba"),
    ("Afeitado Cl\u00e1sico",      45,  25, "\ud83e\ude92",  "Afeitado con navaja y toalla caliente"),
    ("Tratamiento Capilar",   90,  60, "\ud83d\udc86",  "Hidrataci\u00f3n y tratamiento del cuero cabelludo"),
    ("Dise\u00f1o / Puntas",       40,  20, "\ud83c\udfa8",  "Dise\u00f1o en el cabello y puntas"),
    ("Tinte / Color",        120,  90, "\ud83c\udfa8",  "Aplicaci\u00f3n de tinte o color"),
]
for i, row_data in enumerate(servicios):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_s.cell(row=3+i, column=j)
        if j == 2:
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
    ws_s.row_dimensions[3+i].height = 20

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 6: CLIENTES
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_c = wb.create_sheet("\ud83d\udc65 Clientes")
ws_c.sheet_view.showGridLines = False
ws_c.freeze_panes = "A3"

merge_title(ws_c, "A1:H1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 BASE DE DATOS DE CLIENTES", size=12)
ws_c.row_dimensions[1].height = 26

col_c = [
    ("Nombre", 26), ("Tel\u00e9fono", 16), ("Correo", 26),
    ("Nacimiento", 14), ("Notas", 34), ("Fecha Registro", 15),
    ("Visitas", 9), ("Total Gastado", 14)
]
for col, (h, w) in enumerate(col_c, 1):
    c = ws_c.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_c.column_dimensions[get_column_letter(col)].width = w
ws_c.row_dimensions[2].height = 20

clientes = [
    ("Juan Garc\u00eda",   "+502 4111-2222", "juan@gmail.com",  "1990-05-15", "Le gusta el degradado",    "2025-01-15", 8,  640.00),
    ("Pedro L\u00f3pez",   "+502 4222-3333", "pedro@gmail.com", "1985-08-22", "Prefiere navaja",           "2025-01-20", 6,  480.00),
    ("Luis Torres",   "+502 4333-4444", "",                "",           "",                           "2025-02-10", 4,  260.00),
    ("Roberto Ruiz",  "+502 4444-5555", "",                "1998-12-03", "Clientela fija los viernes","2025-03-05", 5,  540.00),
]
for i, row_data in enumerate(clientes):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_c.cell(row=3+i, column=j)
        if j == 8:
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
    ws_c.row_dimensions[3+i].height = 20

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 7: BARBEROS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_b = wb.create_sheet("\ud83d\udc88 Barberos")
ws_b.sheet_view.showGridLines = False
ws_b.freeze_panes = "A3"

merge_title(ws_b, "A1:F1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 EQUIPO DE BARBEROS", size=12)
ws_b.row_dimensions[1].height = 26

col_b = [("Nombre", 24), ("Especialidad", 26), ("Tel\u00e9fono", 16), ("Comisi\u00f3n %", 12), ("Ventas Mes", 14), ("Estado", 10)]
for col, (h, w) in enumerate(col_b, 1):
    c = ws_b.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_b.column_dimensions[get_column_letter(col)].width = w
ws_b.row_dimensions[2].height = 20

barberos = [
    ("Carlos P\u00e9rez",     "Degradados y Fades",  "+502 5555-1111", 40, 2400.00, "\u2705 Activo"),
    ("Jos\u00e9 Mart\u00ednez",    "Cortes Cl\u00e1sicos",      "+502 5555-2222", 35, 1800.00, "\u2705 Activo"),
    ("Miguel Rodr\u00edguez", "Dise\u00f1os y Tintes",     "+502 5555-3333", 40, 2100.00, "\u2705 Activo"),
]
for i, row_data in enumerate(barberos):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_b.cell(row=3+i, column=j)
        if j == 5:
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
        if j == 6:
            c.fill = fill("EAFAF1")
            c.font = Font(name="Calibri", size=10, color=SUCCESS, bold=True)
    ws_b.row_dimensions[3+i].height = 22

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 8: RESUMEN MENSUAL CON GR\u00c1FICA
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_r = wb.create_sheet("\ud83d\udcc8 Resumen Mensual")
ws_r.sheet_view.showGridLines = False

merge_title(ws_r, "A1:D1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 GANANCIAS MENSUALES 2025", size=12)
ws_r.row_dimensions[1].height = 28

for col, (h, w) in enumerate([("Mes", 14), ("Ingresos (Q)", 16), ("# Ventas", 10), ("Promedio", 14)], 1):
    c = ws_r.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_r.column_dimensions[get_column_letter(col)].width = w
ws_r.row_dimensions[2].height = 22

meses_data = [
    ("Enero 2025",     3200.00, 42, 76.19),
    ("Febrero 2025",   4100.00, 55, 74.55),
    ("Marzo 2025",     3800.00, 49, 77.55),
    ("Abril 2025",     5200.00, 68, 76.47),
    ("Mayo 2025",      4900.00, 63, 77.78),
    ("Junio 2025",     6100.00, 78, 78.21),
    ("Julio 2025",        0.00,  0,  0.00),
    ("Agosto 2025",       0.00,  0,  0.00),
    ("Septiembre 2025",   0.00,  0,  0.00),
    ("Octubre 2025",      0.00,  0,  0.00),
    ("Noviembre 2025",    0.00,  0,  0.00),
    ("Diciembre 2025",    0.00,  0,  0.00),
]
for i, row_data in enumerate(meses_data):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_r.cell(row=3+i, column=j)
        if j in (2, 4):
            style_money(c, bg)
            c.value = val
        else:
            style_data(c, bg)
            c.value = val
    ws_r.row_dimensions[3+i].height = 20

# Fila total
tr = 3 + len(meses_data) + 1
ws_r.merge_cells(f"A{tr}:C{tr}")
c_t = ws_r.cell(row=tr, column=1)
c_t.value = "TOTAL A\u00d1O"
c_t.fill = fill(DARK_BG)
c_t.font = Font(name="Calibri", bold=True, size=11, color=GOLD)
c_t.alignment = align("right")
c_total2 = ws_r.cell(row=tr, column=2)
style_total(c_total2)
c_total2.value = sum(r[1] for r in meses_data)

# Gr\u00e1fica de barras
chart = BarChart()
chart.type = "col"
chart.title = "Ganancias Mensuales \u2014 Barber\u00eda El Colocho"
chart.y_axis.title = "Ingresos (Q)"
chart.x_axis.title = "Mes"
chart.style = 10
chart.width = 22
chart.height = 14
chart.grouping = "clustered"

data_ref = Reference(ws_r, min_col=2, min_row=2, max_row=14)
cats_ref = Reference(ws_r, min_col=1, min_row=3, max_row=14)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = GOLD
chart.series[0].graphicalProperties.line.solidFill = GOLD

ws_r.add_chart(chart, "F2")

# Gr\u00e1fica de l\u00ednea overlay
line_chart = LineChart()
line_chart.title = "Tendencia de Ventas"
line_chart.style = 10
line_chart.y_axis.title = "Ventas (Q)"
line_chart.width = 22
line_chart.height = 12

data_ref2 = Reference(ws_r, min_col=2, min_row=2, max_row=8)
cats_ref2 = Reference(ws_r, min_col=1, min_row=3, max_row=8)
line_chart.add_data(data_ref2, titles_from_data=True)
line_chart.set_categories(cats_ref2)
line_chart.series[0].graphicalProperties.line.solidFill = DARK_BG
line_chart.series[0].graphicalProperties.line.width = 25000
ws_r.add_chart(line_chart, "F22")

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 9: REGISTRO MANUAL DE VENTAS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_reg = wb.create_sheet("\ud83d\udcb0 Registro Ventas")
ws_reg.sheet_view.showGridLines = False
ws_reg.freeze_panes = "A4"

merge_title(ws_reg, "A1:L1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 INGRESO MANUAL DE VENTAS", size=12)
ws_reg.row_dimensions[1].height = 28

# Instrucciones
ws_reg.merge_cells("A2:L2")
c_inst = ws_reg["A2"]
c_inst.value = "\ud83d\udcdd Complete cada campo. Las columnas con * son obligatorias. Los totales se calculan autom\u00e1ticamente."
c_inst.fill = fill("FEF9E7")
c_inst.font = Font(name="Calibri", size=10, color=WARNING, italic=True)
c_inst.alignment = align("center")
ws_reg.row_dimensions[2].height = 20

col_reg = [
    ("#*", 6), ("Fecha*", 13), ("Hora*", 9), ("Cliente*", 22), ("Barbero", 18),
    ("Servicio/Producto*", 30), ("Precio Unit.*", 14), ("Cantidad*", 10),
    ("Subtotal", 13), ("Desc%", 8), ("IVA%", 7), ("Total", 13)
]
for col, (h, w) in enumerate(col_reg, 1):
    c = ws_reg.cell(row=3, column=col)
    style_header(c)
    c.value = h
    ws_reg.column_dimensions[get_column_letter(col)].width = w
ws_reg.row_dimensions[3].height = 22

# 20 filas vac\u00edas con f\u00f3rmulas
for i in range(1, 21):
    row = 3 + i
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j in range(1, 13):
        c = ws_reg.cell(row=row, column=j)
        if j == 9:
            c.value = f"=IF(G{row}*H{row}>0,G{row}*H{row},\"\")"
            style_money(c, bg)
        elif j == 12:
            c.value = f"=IF(I{row}<>\"\",I{row}*(1-J{row}/100)*(1+K{row}/100),\"\")"
            style_total(c)
            c.fill = fill("FEF9E7")
            c.font = Font(name="Calibri", bold=True, size=10, color="9A7A2E")
        elif j in (7,):
            style_money(c, bg)
            c.value = ""
        else:
            style_data(c, bg)
            c.value = ""
    ws_reg.row_dimensions[row].height = 20

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# HOJA 10: CITAS
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
ws_citas = wb.create_sheet("\ud83d\udcc5 Citas")
ws_citas.sheet_view.showGridLines = False
ws_citas.freeze_panes = "A3"

merge_title(ws_citas, "A1:G1", "\u2702\ufe0f  BARBER\u00cdA EL COLOCHO \u2014 AGENDA DE CITAS", size=12)
ws_citas.row_dimensions[1].height = 26

col_citas = [("Cliente", 24), ("Barbero", 18), ("Fecha", 13), ("Hora", 9), ("Servicio", 22), ("Estado", 14), ("Notas", 32)]
for col, (h, w) in enumerate(col_citas, 1):
    c = ws_citas.cell(row=2, column=col)
    style_header(c)
    c.value = h
    ws_citas.column_dimensions[get_column_letter(col)].width = w
ws_citas.row_dimensions[2].height = 20

citas_ejemplo = [
    ("Juan Garc\u00eda",   "Carlos P\u00e9rez",   "2025-06-17", "09:00", "Corte Cl\u00e1sico",   "\u2705 Confirmada", ""),
    ("Pedro L\u00f3pez",   "Jos\u00e9 Mart\u00ednez",  "2025-06-17", "10:30", "Corte + Barba",   "\u23f3 Pendiente",  "Llegar puntual"),
    ("Luis Torres",   "Carlos P\u00e9rez",   "2025-06-17", "12:00", "Degradado/Fade",  "\u23f3 Pendiente",  ""),
    ("Roberto Ruiz",  "Miguel R.",      "2025-06-18", "09:00", "Tinte/Color",     "\u2705 Confirmada", "Tinte oscuro"),
]
for i, row_data in enumerate(citas_ejemplo):
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    for j, val in enumerate(row_data, 1):
        c = ws_citas.cell(row=3+i, column=j)
        style_data(c, bg)
        c.value = val
        if j == 6:
            if "Confirmada" in str(val):
                c.fill = fill("EAFAF1")
                c.font = Font(name="Calibri", size=10, color=SUCCESS, bold=True)
            elif "Pendiente" in str(val):
                c.fill = fill("FEF9E7")
                c.font = Font(name="Calibri", size=10, color=WARNING, bold=True)
            elif "Cancelada" in str(val):
                c.fill = fill("FDEDEC")
                c.font = Font(name="Calibri", size=10, color=DANGER, bold=True)
    ws_citas.row_dimensions[3+i].height = 20

# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
# GUARDAR
# \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500
output_path = os.path.join(os.path.dirname(__file__), "barberia_el_colocho_plantilla.xlsx")
wb.save(output_path)
print(f"\
\u2705 Plantilla Excel creada exitosamente:")
print(f"   \ud83d\udcc1 {output_path}")
print(f"\
\ud83d\udccb Hojas incluidas:")
for ws in wb.worksheets:
    print(f"   \u2022 {ws.title}")
print("\
\u2702\ufe0f  Barber\u00eda El Colocho \u2014 Plantilla lista para usar!\
")
